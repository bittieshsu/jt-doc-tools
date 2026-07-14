"""匯出乘車證明清單 — CSV / XLSX / ODS / JSON / XML / TXT / MD。

與電子發票匯出對齊；差別：
- `seq` 序號、`route` 來源-目的（origin → destination 組合）為 computed 欄位。
- JSON 永遠用內部標準格式（ISO 日期 / 整數金額），忽略 field_formats。
"""
from __future__ import annotations

import csv
import io
import json
from datetime import datetime
from typing import Optional

from .settings import FIELD_DEFINITIONS, apply_format, _AMOUNT_FIELD_IDS

_FIELD_DEF_BY_ID = {f["id"]: f for f in FIELD_DEFINITIONS}


def _amount_raw(value):
    """匯出用金額：一律純數字（不加千分位逗號，避免破壞 CSV / 會計軟體匯入）。
    回 int（可存成試算表數值格）或 ""（缺值）。"""
    if value is None or value == "":
        return ""
    try:
        return int(value)
    except (ValueError, TypeError):
        return value


def _resolve_columns(visible_columns: Optional[list[str]],
                     column_order: Optional[list[str]]) -> list[str]:
    if not visible_columns:
        visible_columns = [f["id"] for f in FIELD_DEFINITIONS if f["default_visible"]]
    if not column_order:
        column_order = [f["id"] for f in FIELD_DEFINITIONS]
    return [c for c in column_order if c in visible_columns and c in _FIELD_DEF_BY_ID]


def _label(field_id: str, export_labels: Optional[dict] = None) -> str:
    if export_labels and isinstance(export_labels, dict):
        v = export_labels.get(field_id)
        if isinstance(v, str) and v.strip():
            return v.strip()
    d = _FIELD_DEF_BY_ID.get(field_id)
    return d["label"] if d else field_id


def _route_str(entry: dict) -> str:
    o = (entry.get("origin") or "").strip()
    d = (entry.get("destination") or "").strip()
    return f"{o} → {d}" if (o and d) else (o or d)


def _transport_route_str(entry: dict) -> str:
    t = (entry.get("transport") or "").strip()
    r = _route_str(entry)
    return f"{t} {r}".strip() if (t or r) else ""


def _row_value(entry: dict, field_id: str, row_index: int, field_formats: dict):
    if field_id == "seq":
        return row_index + 1
    if field_id == "route":
        return _route_str(entry)
    if field_id == "transport_route":
        return _transport_route_str(entry)
    if field_id in _AMOUNT_FIELD_IDS:
        return _amount_raw(entry.get(field_id))   # 匯出金額一律純數字（無逗號）
    return apply_format(field_id, entry.get(field_id), field_formats)


def _raw_value(entry: dict, field_id: str, row_index: int):
    """JSON 用：computed 欄位算出，其餘取內部原值。"""
    if field_id == "seq":
        return row_index + 1
    if field_id == "route":
        return _route_str(entry)
    if field_id == "transport_route":
        return _transport_route_str(entry)
    return entry.get(field_id)


def export_csv(entries, columns, field_formats, export_labels=None) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, lineterminator="\n")
    w.writerow([_label(c, export_labels) for c in columns])
    for i, e in enumerate(entries):
        w.writerow([_row_value(e, c, i, field_formats) for c in columns])
    return ("﻿" + buf.getvalue()).encode("utf-8")   # BOM for Excel CJK


def export_xlsx(entries, columns, field_formats, export_labels=None) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise RuntimeError("openpyxl 未安裝，無法匯出 xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "transit"
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="2563EB")
    ha = Alignment(horizontal="center", vertical="center")
    for ci, c in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=ci, value=_label(c, export_labels))
        cell.font = hf
        cell.fill = hfill
        cell.alignment = ha
        ws.column_dimensions[cell.column_letter].width = 40 if c == "note" else (
            22 if c in ("route", "train", "ticket_no") else 14)
    for i, e in enumerate(entries):
        for ci, c in enumerate(columns, start=1):
            v = _row_value(e, c, i, field_formats)
            if c == "seq":
                ws.cell(row=i + 2, column=ci, value=int(v))
            else:
                ws.cell(row=i + 2, column=ci, value=v if v != "" else None)
    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_json(entries, columns, field_formats) -> bytes:
    out = []
    for i, e in enumerate(entries):
        out.append({c: _raw_value(e, c, i) for c in columns})
    return json.dumps({"entries": out, "exported_at": datetime.now().isoformat()},
                      ensure_ascii=False, indent=2).encode("utf-8")


def export_ods(entries, columns, field_formats, export_labels=None) -> bytes:
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.style import Style, TextProperties, TableCellProperties
        from odf.table import Table, TableColumn, TableRow, TableCell
        from odf.text import P
    except ImportError:
        raise RuntimeError("odfpy 未安裝，無法匯出 ods")
    doc = OpenDocumentSpreadsheet()
    hdr = Style(name="HdrCell", family="table-cell")
    hdr.addElement(TextProperties(fontweight="bold", color="#FFFFFF"))
    hdr.addElement(TableCellProperties(backgroundcolor="#2563EB"))
    doc.styles.addElement(hdr)
    table = Table(name="transit")
    for _ in columns:
        table.addElement(TableColumn())
    hr = TableRow()
    for c in columns:
        cell = TableCell(stylename=hdr)
        cell.addElement(P(text=_label(c, export_labels)))
        hr.addElement(cell)
    table.addElement(hr)
    for i, e in enumerate(entries):
        tr = TableRow()
        for c in columns:
            v = _row_value(e, c, i, field_formats)
            cell = TableCell()
            cell.addElement(P(text=str(v) if v is not None else ""))
            tr.addElement(cell)
        table.addElement(tr)
    doc.spreadsheet.addElement(table)
    buf = io.BytesIO()
    doc.write(buf)
    return buf.getvalue()


def _xml_escape(s):
    if s is None:
        return ""
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;").replace("'", "&apos;"))


def export_xml(entries, columns, field_formats) -> bytes:
    lines = ['<?xml version="1.0" encoding="utf-8"?>', "<transit_proofs>"]
    for i, e in enumerate(entries):
        lines.append("  <entry>")
        for c in columns:
            lines.append(f"    <{c}>{_xml_escape(_row_value(e, c, i, field_formats))}</{c}>")
        lines.append("  </entry>")
    lines.append("</transit_proofs>")
    return ("\n".join(lines) + "\n").encode("utf-8")


def _display_width(s):
    if s is None:
        return 0
    n = 0
    for ch in str(s):
        cp = ord(ch)
        n += 2 if ((0x3000 <= cp <= 0x9FFF) or (0xFF00 <= cp <= 0xFFEF)) else 1
    return n


def _pad(s, width):
    s = str(s) if s is not None else ""
    return s + " " * max(0, width - _display_width(s))


def export_txt(entries, columns, field_formats, export_labels=None) -> bytes:
    headers = [_label(c, export_labels) for c in columns]
    widths = [_display_width(h) for h in headers]
    rows = []
    for i, e in enumerate(entries):
        row = [str(_row_value(e, c, i, field_formats)) for c in columns]
        rows.append(row)
        for j, v in enumerate(row):
            widths[j] = max(widths[j], _display_width(v))
    out = ["  ".join(_pad(headers[j], widths[j]) for j in range(len(columns))),
           "  ".join("-" * widths[j] for j in range(len(columns)))]
    for row in rows:
        out.append("  ".join(_pad(row[j], widths[j]) for j in range(len(columns))))
    return ("\n".join(out) + "\n").encode("utf-8")


def _md_escape(s):
    if s is None:
        return ""
    return str(s).replace("|", "\\|").replace("\n", " ").replace("\r", "")


def export_md(entries, columns, field_formats, export_labels=None) -> bytes:
    headers = [_label(c, export_labels) for c in columns]
    out = ["| " + " | ".join(_md_escape(h) for h in headers) + " |",
           "| " + " | ".join("---" for _ in columns) + " |"]
    for i, e in enumerate(entries):
        out.append("| " + " | ".join(
            _md_escape(_row_value(e, c, i, field_formats)) for c in columns) + " |")
    return ("\n".join(out) + "\n").encode("utf-8")


def build_export(entries: list[dict], visible_columns: list[str],
                 column_order: list[str], field_formats: dict, fmt: str,
                 export_labels: Optional[dict] = None) -> tuple[bytes, str, str]:
    columns = _resolve_columns(visible_columns, column_order)
    fmt = (fmt or "").lower()
    if fmt == "csv":
        return (export_csv(entries, columns, field_formats, export_labels),
                "text/csv; charset=utf-8", "csv")
    if fmt == "xlsx":
        return (export_xlsx(entries, columns, field_formats, export_labels),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx")
    if fmt == "ods":
        return (export_ods(entries, columns, field_formats, export_labels),
                "application/vnd.oasis.opendocument.spreadsheet", "ods")
    if fmt == "json":
        return (export_json(entries, columns, field_formats),
                "application/json; charset=utf-8", "json")
    if fmt == "xml":
        return (export_xml(entries, columns, field_formats),
                "application/xml; charset=utf-8", "xml")
    if fmt == "txt":
        return (export_txt(entries, columns, field_formats, export_labels),
                "text/plain; charset=utf-8", "txt")
    if fmt == "md":
        return (export_md(entries, columns, field_formats, export_labels),
                "text/markdown; charset=utf-8", "md")
    raise ValueError(f"不支援的匯出格式：{fmt}")
