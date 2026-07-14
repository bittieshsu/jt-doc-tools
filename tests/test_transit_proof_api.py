"""乘車證明工具端點整合測試（合成 PDF，auth OFF = 單機）。

用 PyMuPDF 產生符合版面的假憑證 PDF，走 upload → buffer → 匯出 → 設定 全流程。
"""
from __future__ import annotations

import fitz
import pytest


def _make_pdf(text: str) -> bytes:
    # 用 PyMuPDF 內建 CJK 字型（預設 helv 無中文 glyph，抽出會變 notdef）。
    # 逐行插入避免 insert_text 的換行處理差異。
    doc = fitz.open()
    page = doc.new_page(width=595, height=842)
    y = 60
    for line in text.split("\n"):
        page.insert_text((50, y), line, fontsize=11, fontname="china-t")
        y += 18
    return doc.tobytes()


_HSR_TEXT = ("統一編號 : 00000000\n台灣高鐵電子車票證明\n卡號/票號 : 0000000000000\n"
             "乘車日期 : 2026-06-09 07:20:00\n起程站 : 高鐵AAA車站\n到達站 : 高鐵BBB車站\n"
             "銷售額 : 700\n營業稅額 : 35\n票價 : 735\n")


@pytest.fixture(autouse=True)
def _clear(client):
    client.delete("/tools/transit-proof/buffer")
    yield
    client.delete("/tools/transit-proof/buffer")


def test_page_renders(client):
    r = client.get("/tools/transit-proof/")
    assert r.status_code == 200
    assert "乘車證明整理" in r.text


def test_upload_parse_and_buffer(client):
    pdf = _make_pdf(_HSR_TEXT)
    r = client.post("/tools/transit-proof/upload",
                    files=[("files", ("hsr.pdf", pdf, "application/pdf"))])
    assert r.status_code == 200, r.text
    j = r.json()
    assert j["added"] == 1
    assert len(j["entries"]) == 1
    e = j["entries"][0]
    assert e["transport"] == "高鐵"
    assert e["origin"] == "AAA" and e["destination"] == "BBB"
    assert e["fare"] == 735

    r2 = client.post("/tools/transit-proof/upload",
                     files=[("files", ("hsr.pdf", pdf, "application/pdf"))])
    assert r2.json()["duplicates"] == 1
    assert r2.json()["added"] == 0


def test_upload_rejects_non_transit_pdf(client):
    pdf = _make_pdf("這是一份普通報告，不是乘車證明。")
    r = client.post("/tools/transit-proof/upload",
                    files=[("files", ("x.pdf", pdf, "application/pdf"))])
    assert r.status_code == 200
    j = r.json()
    assert j["added"] == 0
    assert len(j["failed"]) == 1


def test_export_all_formats(client):
    pdf = _make_pdf(_HSR_TEXT)
    client.post("/tools/transit-proof/upload",
                files=[("files", ("hsr.pdf", pdf, "application/pdf"))])
    for fmt in ("csv", "xlsx", "ods", "json", "xml", "txt", "md"):
        r = client.post("/tools/transit-proof/export", json={"format": fmt})
        assert r.status_code == 200, f"{fmt}: {r.text}"
        assert len(r.content) > 0
    r = client.post("/tools/transit-proof/export", json={"format": "pdf"})
    assert r.status_code == 400


def test_export_empty_buffer_400(client):
    r = client.post("/tools/transit-proof/export", json={"format": "csv"})
    assert r.status_code == 400


def test_csv_default_columns(client):
    pdf = _make_pdf(_HSR_TEXT)
    client.post("/tools/transit-proof/upload",
                files=[("files", ("hsr.pdf", pdf, "application/pdf"))])
    r = client.post("/tools/transit-proof/export", json={"format": "csv"})
    text = r.content.decode("utf-8-sig")
    header = text.splitlines()[0]
    assert header == "日期,交通工具,來源-目的,費用"   # 預設 4 欄
    assert "高鐵" in text and "AAA → BBB" in text and "735" in text


def test_settings_roundtrip_and_applied(client):
    pdf = _make_pdf(_HSR_TEXT)
    client.post("/tools/transit-proof/upload",
                files=[("files", ("hsr.pdf", pdf, "application/pdf"))])
    r = client.post("/tools/transit-proof/settings", json={
        "visible_columns": ["date", "fare", "ticket_no"],
        "column_order": ["date", "fare", "ticket_no"],
        "field_formats": {"fare": "comma", "date": "roc"},
        "export_labels": {"fare": "票價(元)"},
    })
    assert r.status_code == 200
    csv = client.post("/tools/transit-proof/export", json={"format": "csv"}).content.decode("utf-8-sig")
    header = csv.splitlines()[0]
    assert header == "日期,票價(元),票號 / 卡號"
    assert "115/06/09" in csv       # 民國格式套用

    client.post("/tools/transit-proof/settings", json={
        "visible_columns": ["date", "transport", "route", "fare"],
        "column_order": [], "field_formats": {}, "export_labels": {}})


def test_delete_entry(client):
    pdf = _make_pdf(_HSR_TEXT)
    up = client.post("/tools/transit-proof/upload",
                     files=[("files", ("hsr.pdf", pdf, "application/pdf"))]).json()
    eid = up["entries"][0]["id"]
    r = client.delete(f"/tools/transit-proof/entry/{eid}")
    assert r.status_code == 200
    assert client.get("/tools/transit-proof/buffer").json()["entries"] == []


_HSR_1050 = _HSR_TEXT.replace("票價 : 735", "票價 : 1050").replace("銷售額 : 700", "銷售額 : 1000").replace("營業稅額 : 35", "營業稅額 : 50")


def test_export_amounts_have_no_comma(client):
    """匯出費用 / 金額欄一律純數字（無千分位逗號），即使顯示格式設 comma
    （v1.12.77）。"""
    client.post("/tools/transit-proof/upload",
                files=[("files", ("h.pdf", _make_pdf(_HSR_1050), "application/pdf"))])
    client.post("/tools/transit-proof/settings", json={
        "visible_columns": ["date", "fare", "amount_untaxed", "tax"],
        "column_order": ["date", "fare", "amount_untaxed", "tax"],
        "field_formats": {"fare": "comma", "amount_untaxed": "comma", "tax": "comma"},
    })
    csv = client.post("/tools/transit-proof/export", json={"format": "csv"}).content.decode("utf-8-sig")
    assert "1,050" not in csv and "1050" in csv
    # XLSX 金額存數值格
    import openpyxl, io as _io
    x = client.post("/tools/transit-proof/export", json={"format": "xlsx"}).content
    ws = openpyxl.load_workbook(_io.BytesIO(x)).active
    vals = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    assert 1050 in vals and all(not (isinstance(v, str) and "," in v) for v in vals if v is not None)
    client.post("/tools/transit-proof/settings", json={
        "visible_columns": ["date", "transport", "route", "fare"],
        "column_order": [], "field_formats": {}, "export_labels": {}})


def test_export_transport_route_combined_field(client):
    """合併欄位「交通工具-來源-目的」= 交通工具 + 起訖（v1.12.75）。"""
    client.post("/tools/transit-proof/upload",
                files=[("files", ("h.pdf", _make_pdf(_HSR_TEXT), "application/pdf"))])
    client.post("/tools/transit-proof/settings", json={
        "visible_columns": ["transport_route", "fare"],
        "column_order": ["transport_route", "fare"]})
    csv = client.post("/tools/transit-proof/export", json={"format": "csv"}).content.decode("utf-8-sig")
    header = csv.splitlines()[0]
    assert header.split(",")[0] == "交通工具-來源-目的"
    assert "高鐵 AAA → BBB" in csv
    client.post("/tools/transit-proof/settings", json={
        "visible_columns": ["date", "transport", "route", "fare"],
        "column_order": [], "field_formats": {}, "export_labels": {}})


def test_delete_batch(client):
    # 造 3 筆（改票號避免去重）
    for tk in ("1111111111111", "2222222222222", "3333333333333"):
        txt = _HSR_TEXT.replace("0000000000000", tk)
        client.post("/tools/transit-proof/upload",
                    files=[("files", ("h.pdf", _make_pdf(txt), "application/pdf"))])
    entries = client.get("/tools/transit-proof/buffer").json()["entries"]
    assert len(entries) == 3
    ids = [entries[0]["id"], entries[1]["id"]]
    r = client.post("/tools/transit-proof/buffer/delete-batch", json={"ids": ids})
    assert r.status_code == 200, r.text
    j = r.json()
    assert j["deleted"] == 2
    assert len(j["entries"]) == 1
    # ids 非陣列 → 400
    assert client.post("/tools/transit-proof/buffer/delete-batch", json={"ids": "x"}).status_code == 400


def test_page_ui_structure(client):
    """鎖住乘車證明頁的關鍵 UI 標記，抓「元素 / 功能被移除」的 regression。
    純伺服器端渲染 + inline JS 斷言（不需瀏覽器）。"""
    html = client.get("/tools/transit-proof/").text
    # 批次勾選：全選 / 列勾選 / 刪除選取鈕
    assert 'id="tpSelAll"' in html or "tpSelAll" in html
    assert "tp-rowchk" in html
    assert 'id="tpBtnDelSel"' in html and "刪除選取" in html
    # 列刪除：垃圾桶 icon 來源 + 刪除前確認
    assert 'id="tpIconTrash"' in html
    assert "showConfirm" in html and "刪除乘車證明" in html
    # 「從工作區載入」對齊電子發票頁（ws-load-btn + 共用 helper）
    assert "ws-load-btn" in html and "attachWorkspaceLoadButton" in html
    # 工具列「欄位」按鈕（非「設定」）
    assert "欄位" in html
    # 表頭可排序 + 卡片可收折（h2 面板標題）
    assert "tp-sortable" in html and "tp-panel-h2" in html
    # 欄位設定拖曳排序（grab 把手，非 ▲▼）
    assert "tp-fr-grab" in html
    # 合併欄位選項
    assert "transport_route" in html or "交通工具-來源-目的" in html


def test_api_endpoint_no_buffer(client):
    pdf = _make_pdf(_HSR_TEXT)
    r = client.post("/tools/transit-proof/api/transit-proof",
                    files=[("files", ("hsr.pdf", pdf, "application/pdf"))])
    assert r.status_code == 200
    j = r.json()
    assert j["count"] == 1
    assert j["entries"][0]["transport"] == "高鐵"
    assert client.get("/tools/transit-proof/buffer").json()["entries"] == []
